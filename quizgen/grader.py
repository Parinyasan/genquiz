from openpyxl import load_workbook
import numpy as np
import os
import pickle
from .generator import *


def Grading(answer_path, solution_path, csv_file='grading.csv'):
    csv = open(csv_file, 'w')
    for fn in os.listdir(answer_path):
        if fn.lower().endswith('.xlsx'):
            student_id = fn[:-5]
            sol_fn = os.path.join(solution_path, student_id + '.pk')
            if os.path.exists(sol_fn):
                solutions = pickle.load(open(sol_fn, 'rb'))
                wb = load_workbook(os.path.join(answer_path, fn))
                ws = wb.active
                csv.write(student_id)
                for solution in solutions:
                    ans = ws[solution.excel_cells]
                    if isinstance(ans, tuple):
                        ans = np.array([[i.value for i in j] for j in ans])
                    else:
                        ans = np.array([ans.value])
                    score = np.sum(ans.astype(np.str) == solution.solution.astype(np.str)) / ans.size * solution.score
                    csv.write(',' + str(score))
                csv.write('\n')
    csv.close()
