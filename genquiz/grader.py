from openpyxl import load_workbook
import numpy as np
import os
import pickle
from .generator import *


def is_numeric(obj):
    attrs = [np.int, np.int32, np.float, np.float32, np.float64]
    return any(obj.dtype == attr for attr in attrs)


def Grading(answer_path, solution_path, csv_file='grading.csv'):
    csv = open(csv_file, 'w')
    for fn in os.listdir(answer_path):
        if fn.lower().endswith('.xlsx'):
            student_id = fn[:-5]
            sol_fn = os.path.join(solution_path, student_id + '.pk')
            if os.path.exists(sol_fn):
                solutions = pickle.load(open(sol_fn, 'rb'))
                wb = load_workbook(os.path.join(answer_path, fn))
                ws = wb.active
                csv.write(student_id)
                for solution in solutions:
                    ans = ws[solution.excel_cells]
                    if isinstance(ans, tuple):
                        ans = np.array([[i.value for i in j] for j in ans])
                    else:
                        ans = np.array([ans.value])
                    if is_numeric(ans):
                        score = np.abs(ans - solution.solution)
                        score = np.sum(score <= 0.1) / ans.size * solution.score
                    else:
                        score = np.sum(ans.astype(np.str) == solution.solution.astype(np.str)) / ans.size * solution.score
                    csv.write(',' + str(score))
                csv.write('\n')
    csv.close()
